from copy import copy
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


EXPORT_HEADERS = [
    "Validation Status",
    "Evidence Assessment",
    "Supporting Evidence",
    "Reviewer Action",
]


EXPORT_COLUMN_WIDTHS = {
    "Validation Status": 22,
    "Evidence Assessment": 65,
    "Supporting Evidence": 55,
    "Reviewer Action": 45,
}


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F2937",
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)


THIN_BORDER_SIDE = Side(
    style="thin",
    color="D1D5DB",
)

OUTPUT_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)


STATUS_STYLES = {
    "validated": {
        "fill": "E2F0D9",
        "font": "375623",
    },
    "partially validated": {
        "fill": "FFF2CC",
        "font": "7F6000",
    },
    "not validated": {
        "fill": "FCE4D6",
        "font": "9C5700",
    },
    "contradicted": {
        "fill": "F4CCCC",
        "font": "9C0006",
    },
    "insufficient evidence": {
        "fill": "E7E6E6",
        "font": "44546A",
    },
    "error": {
        "fill": "F4CCCC",
        "font": "9C0006",
    },
}


def get_validated_workbook_path(questionnaire_path):
    questionnaire_path = Path(questionnaire_path)

    return questionnaire_path.with_name(
        f"{questionnaire_path.stem}_validated"
        f"{questionnaire_path.suffix}"
    )


def _should_reset_validated_copy(
    questionnaire_path,
    validated_path,
):
    if not validated_path.exists():
        return True

    return (
        questionnaire_path.stat().st_mtime_ns
        > validated_path.stat().st_mtime_ns
    )


def _copy_cell_style(
    source_cell,
    target_cell,
):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(
            source_cell.alignment
        )
        target_cell.number_format = (
            source_cell.number_format
        )
        target_cell.protection = copy(
            source_cell.protection
        )


def _find_or_create_export_columns(
    worksheet,
    header_row,
):
    existing_headers = {}

    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        value = worksheet.cell(
            row=header_row,
            column=column_number,
        ).value

        if value is None:
            continue

        existing_headers[
            str(value).strip()
        ] = column_number

    column_map = {}

    next_column = worksheet.max_column + 1

    style_source = None

    for column_number in range(
        worksheet.max_column,
        0,
        -1,
    ):
        candidate = worksheet.cell(
            row=header_row,
            column=column_number,
        )

        if candidate.value is not None:
            style_source = candidate
            break

    for header in EXPORT_HEADERS:
        if header in existing_headers:
            column_map[header] = (
                existing_headers[header]
            )
            continue

        target_cell = worksheet.cell(
            row=header_row,
            column=next_column,
            value=header,
        )

        if style_source is not None:
            _copy_cell_style(
                style_source,
                target_cell,
            )

        column_map[header] = next_column

        next_column += 1

    return column_map


def _style_export_headers(
    worksheet,
    header_row,
    column_map,
):
    for column_number in column_map.values():
        cell = worksheet.cell(
            row=header_row,
            column=column_number,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = OUTPUT_BORDER

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _style_output_cell(
    cell,
):
    cell.border = OUTPUT_BORDER

    cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )


def _style_status_cell(
    cell,
    status,
):
    status_key = (
        str(status).strip().lower()
        if status
        else ""
    )

    style = STATUS_STYLES.get(
        status_key
    )

    cell.border = OUTPUT_BORDER

    cell.alignment = Alignment(
        horizontal="center",
        vertical="top",
        wrap_text=True,
    )

    if style:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=style["fill"],
        )

        cell.font = Font(
            bold=True,
            color=style["font"],
        )
    else:
        cell.font = Font(
            bold=True,
        )


def _humanize_status(status):
    if not status:
        return ""

    return str(status).replace(
        "_",
        " ",
    ).title()


def _format_list_section(
    label,
    values,
):
    if not values:
        return None

    cleaned_values = [
        str(value).strip()
        for value in values
        if value is not None
        and str(value).strip()
    ]

    if not cleaned_values:
        return None

    lines = [
        f"{label}:",
    ]

    lines.extend(
        f"• {value}"
        for value in cleaned_values
    )

    return "\n".join(lines)


def _build_evidence_assessment(validation):
    sections = []

    confidence = validation.get(
        "confidence"
    )

    evidence_strength = validation.get(
        "evidence_strength"
    )

    metadata = []

    if confidence:
        metadata.append(
            f"Confidence: "
            f"{str(confidence).title()}"
        )

    if evidence_strength:
        metadata.append(
            f"Evidence strength: "
            f"{str(evidence_strength).title()}"
        )

    if metadata:
        sections.append(
            " | ".join(metadata)
        )

    explanation = validation.get(
        "explanation"
    )

    if explanation:
        sections.append(
            str(explanation).strip()
        )

    list_sections = [
        (
            "Evidence supports",
            validation.get(
                "evidence_proves"
            ),
        ),
        (
            "Evidence does not establish",
            validation.get(
                "evidence_does_not_prove"
            ),
        ),
        (
            "Gaps",
            validation.get(
                "gaps"
            ),
        ),
        (
            "Contradictions",
            validation.get(
                "contradictions"
            ),
        ),
        (
            "Additional evidence needed",
            validation.get(
                "additional_evidence_needed"
            ),
        ),
    ]

    for label, values in list_sections:
        section = _format_list_section(
            label,
            values,
        )

        if section:
            sections.append(section)

    return "\n\n".join(sections)


def _build_supporting_evidence(
    resolved_sources,
):
    if not resolved_sources:
        return ""

    source_lines = []

    for source in resolved_sources:
        citation_label = source.get(
            "citation_label"
        )

        display_name = source.get(
            "display_name"
        )

        file_name = source.get(
            "file_name"
        )

        if citation_label:
            source_lines.append(
                str(citation_label)
            )

        elif display_name:
            source_lines.append(
                str(display_name)
            )

        elif file_name:
            source_lines.append(
                str(file_name)
            )

    return "\n".join(source_lines)


def _write_result_to_row(
    worksheet,
    row_number,
    column_map,
    result,
):
    validation = result.get(
        "validation"
    )

    error = result.get(
        "error"
    )

    if error:
        status = "Error"

        assessment = (
            f"Validation error: {error}"
        )

        supporting_evidence = ""

        reviewer_action = (
            "Review the validation error "
            "and rerun this questionnaire item."
        )

    elif validation:
        status = _humanize_status(
            validation.get(
                "status"
            )
        )

        assessment = (
            _build_evidence_assessment(
                validation
            )
        )

        supporting_evidence = (
            _build_supporting_evidence(
                result.get(
                    "resolved_sources",
                    [],
                )
            )
        )

        reviewer_action = (
            validation.get(
                "reviewer_action"
            )
            or ""
        )

    else:
        status = "Error"

        assessment = (
            "No validation result was available."
        )

        supporting_evidence = ""

        reviewer_action = (
            "Review this questionnaire item "
            "and rerun validation."
        )

    values = {
        "Validation Status": status,
        "Evidence Assessment": assessment,
        "Supporting Evidence": (
            supporting_evidence
        ),
        "Reviewer Action": (
            reviewer_action
        ),
    }

    for header, value in values.items():
        cell = worksheet.cell(
            row=row_number,
            column=column_map[header],
            value=value,
        )

        if header == "Validation Status":
            _style_status_cell(
                cell,
                status,
            )
        else:
            _style_output_cell(cell)


def export_validation_results(
    questionnaire_path,
    validation_results,
    sheet_header_rows,
):
    questionnaire_path = Path(
        questionnaire_path
    )

    if (
        questionnaire_path.suffix.lower()
        != ".xlsx"
    ):
        raise ValueError(
            "Validated questionnaire export "
            "currently supports XLSX files only."
        )

    if not questionnaire_path.exists():
        raise FileNotFoundError(
            questionnaire_path
        )

    validated_path = (
        get_validated_workbook_path(
            questionnaire_path
        )
    )

    if _should_reset_validated_copy(
        questionnaire_path,
        validated_path,
    ):
        shutil.copy2(
            questionnaire_path,
            validated_path,
        )

    workbook = load_workbook(
        validated_path
    )

    column_maps = {}

    for result in validation_results:
        question = result.get(
            "question",
            {},
        )

        sheet_name = question.get(
            "source_sheet"
        )

        source_row = question.get(
            "source_row"
        )

        if (
            not sheet_name
            or not source_row
        ):
            continue

        if sheet_name not in workbook.sheetnames:
            continue

        if sheet_name not in sheet_header_rows:
            continue

        worksheet = workbook[
            sheet_name
        ]

        if sheet_name not in column_maps:
            pandas_header_row = (
                sheet_header_rows[
                    sheet_name
                ]
            )

            excel_header_row = (
                int(pandas_header_row)
                + 1
            )

            column_maps[
                sheet_name
            ] = (
                _find_or_create_export_columns(
                    worksheet,
                    excel_header_row,
                )
            )

            _style_export_headers(
                worksheet=worksheet,
                header_row=excel_header_row,
                column_map=(
                    column_maps[
                        sheet_name
                    ]
                ),
            )

        _write_result_to_row(
            worksheet=worksheet,
            row_number=int(source_row),
            column_map=(
                column_maps[
                    sheet_name
                ]
            ),
            result=result,
        )

    for sheet_name, column_map in (
        column_maps.items()
    ):
        worksheet = workbook[
            sheet_name
        ]

        for header, column_number in (
            column_map.items()
        ):
            column_letter = (
                get_column_letter(
                    column_number
                )
            )

            worksheet.column_dimensions[
                column_letter
            ].width = (
                EXPORT_COLUMN_WIDTHS[
                    header
                ]
            )

    workbook.save(
        validated_path
    )

    return validated_path