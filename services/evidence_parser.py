from pathlib import Path
import csv
import re

import pymupdf
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook


def split_text_into_chunks(
    text,
    chunk_size=1800,
    overlap=200,
):
    chunks = []

    if not text:
        return chunks

    start = 0
    chunk_number = 1

    while start < len(text):
        end = min(
            start + chunk_size,
            len(text),
        )

        chunk_text = text[start:end]

        if end < len(text):
            last_space = chunk_text.rfind(" ")

            if last_space > 0:
                end = start + last_space
                chunk_text = text[start:end]

        chunk_text = chunk_text.strip()

        if chunk_text:
            chunks.append(
                {
                    "chunk_number": chunk_number,
                    "text": chunk_text,
                }
            )

        if end >= len(text):
            break

        start = max(
            end - overlap,
            start + 1,
        )

        chunk_number += 1

    return chunks


# -------------------------
# PDF
# -------------------------

def is_likely_heading(text):
    text = text.strip()

    if not text:
        return False

    if len(text) > 100:
        return False

    if "\n" in text:
        lines = [
            line.strip()
            for line in text.splitlines()
            if line.strip()
        ]

        if len(lines) > 2:
            return False

    if text.endswith(
        (
            ".",
            "?",
            "!",
            ";",
            ":",
        )
    ):
        return False

    if text.startswith(
        (
            "●",
            "•",
            "-",
            "○",
        )
    ):
        return False

    words = re.findall(
        r"[A-Za-z][A-Za-z0-9'-]*",
        text,
    )

    if not words:
        return False

    if len(words) > 12:
        return False

    heading_like_words = sum(
        1
        for word in words
        if (
            word[0].isupper()
            or word.isupper()
        )
    )

    return (
        heading_like_words
        / len(words)
        >= 0.60
    )


def split_pdf_page_into_passages(text):
    if not text:
        return []

    raw_blocks = re.split(
        r"\n\s*\n+",
        text,
    )

    blocks = [
        block.strip()
        for block in raw_blocks
        if block.strip()
    ]

    passages = []

    current_heading = None

    for block in blocks:
        if is_likely_heading(block):
            current_heading = block
            continue

        passage_text = block

        if current_heading:
            passage_text = (
                f"{current_heading}\n"
                f"{block}"
            )

        passages.append(
            {
                "heading": current_heading,
                "text": passage_text,
            }
        )

    if not passages and text.strip():
        passages.append(
            {
                "heading": None,
                "text": text.strip(),
            }
        )

    return passages


def extract_pdf_pages(
    file_path,
    evidence_id,
):
    file_path = Path(file_path)

    pages = []

    with pymupdf.open(file_path) as document:
        for page_index in range(
            document.page_count
        ):
            page = document[page_index]

            text = page.get_text(
                "text",
                sort=True,
            ).strip()

            pages.append(
                {
                    "evidence_id": evidence_id,
                    "file_name": file_path.name,
                    "pdf_page_number": (
                        page_index + 1
                    ),
                    "text": text,
                    "extractable_text": bool(
                        text
                    ),
                }
            )

    return pages


def chunk_pdf_pages(pages):
    chunks = []

    for page in pages:
        passages = (
            split_pdf_page_into_passages(
                page["text"]
            )
        )

        chunk_number = 1

        for passage in passages:
            passage_chunks = (
                split_text_into_chunks(
                    passage["text"],
                    chunk_size=900,
                    overlap=100,
                )
            )

            for passage_chunk in passage_chunks:
                source_id = (
                    f"{page['evidence_id']}"
                    f"-P{page['pdf_page_number']:03d}"
                    f"-C{chunk_number:02d}"
                )

                provenance = {
                    "pdf_page_number": page[
                        "pdf_page_number"
                    ],
                }

                if passage.get(
                    "heading"
                ):
                    provenance[
                        "section_heading"
                    ] = passage["heading"]

                chunks.append(
                    {
                        "source_id": source_id,
                        "evidence_id": page[
                            "evidence_id"
                        ],
                        "file_name": page[
                            "file_name"
                        ],
                        "file_type": ".pdf",
                        "chunk_number": (
                            chunk_number
                        ),
                        "text": passage_chunk[
                            "text"
                        ],
                        "provenance": (
                            provenance
                        ),
                    }
                )

                chunk_number += 1

    return chunks


def process_pdf(
    file_path,
    evidence_id,
):
    file_path = Path(file_path)

    pages = extract_pdf_pages(
        file_path=file_path,
        evidence_id=evidence_id,
    )

    page_count = len(pages)

    extractable_page_count = sum(
        1
        for page in pages
        if page["extractable_text"]
    )

    if extractable_page_count == 0:
        extraction_status = (
            "no_extractable_text"
        )

    elif (
        extractable_page_count
        < page_count
    ):
        extraction_status = "partial"

    else:
        extraction_status = "extracted"

    chunks = chunk_pdf_pages(pages)

    return {
        "evidence_id": evidence_id,
        "file_name": file_path.name,
        "file_type": ".pdf",
        "page_count": page_count,
        "extractable_page_count": (
            extractable_page_count
        ),
        "extraction_status": (
            extraction_status
        ),
        "chunk_count": len(chunks),
        "pages": pages,
        "chunks": chunks,
    }


# -------------------------
# DOCX
# -------------------------

def extract_docx_blocks(
    file_path,
    evidence_id,
):
    file_path = Path(file_path)

    document = Document(file_path)

    blocks = []
    block_number = 1
    table_number = 0
    current_heading = None

    for child in (
        document.element.body.iterchildren()
    ):
        if child.tag.endswith("}p"):
            paragraph = Paragraph(
                child,
                document,
            )

            text = paragraph.text.strip()

            if not text:
                continue

            style_name = (
                paragraph.style.name
                if paragraph.style
                else ""
            )

            if style_name.lower().startswith(
                "heading"
            ):
                current_heading = text

            blocks.append(
                {
                    "evidence_id": (
                        evidence_id
                    ),
                    "file_name": (
                        file_path.name
                    ),
                    "block_number": (
                        block_number
                    ),
                    "block_type": (
                        "paragraph"
                    ),
                    "heading": (
                        current_heading
                    ),
                    "text": text,
                }
            )

            block_number += 1

        elif child.tag.endswith("}tbl"):
            table_number += 1

            table = Table(
                child,
                document,
            )

            for row_number, row in enumerate(
                table.rows,
                start=1,
            ):
                cell_values = []

                for cell in row.cells:
                    cell_text = " ".join(
                        cell.text.split()
                    ).strip()

                    if cell_text:
                        cell_values.append(
                            cell_text
                        )

                row_text = " | ".join(
                    cell_values
                )

                if not row_text:
                    continue

                blocks.append(
                    {
                        "evidence_id": (
                            evidence_id
                        ),
                        "file_name": (
                            file_path.name
                        ),
                        "block_number": (
                            block_number
                        ),
                        "block_type": (
                            "table_row"
                        ),
                        "heading": (
                            current_heading
                        ),
                        "table_number": (
                            table_number
                        ),
                        "row_number": (
                            row_number
                        ),
                        "text": row_text,
                    }
                )

                block_number += 1

    return blocks


def chunk_docx_blocks(blocks):
    chunks = []

    for block in blocks:
        block_chunks = (
            split_text_into_chunks(
                block["text"]
            )
        )

        for chunk in block_chunks:
            source_id = (
                f"{block['evidence_id']}"
                f"-B{block['block_number']:03d}"
                f"-C{chunk['chunk_number']:02d}"
            )

            provenance = {
                "block_number": block[
                    "block_number"
                ],
                "block_type": block[
                    "block_type"
                ],
            }

            if block.get("heading"):
                provenance[
                    "heading"
                ] = block["heading"]

            if (
                block["block_type"]
                == "table_row"
            ):
                provenance[
                    "table_number"
                ] = block["table_number"]

                provenance[
                    "row_number"
                ] = block["row_number"]

            chunks.append(
                {
                    "source_id": (
                        source_id
                    ),
                    "evidence_id": block[
                        "evidence_id"
                    ],
                    "file_name": block[
                        "file_name"
                    ],
                    "file_type": ".docx",
                    "chunk_number": chunk[
                        "chunk_number"
                    ],
                    "text": chunk["text"],
                    "provenance": (
                        provenance
                    ),
                }
            )

    return chunks


def process_docx(
    file_path,
    evidence_id,
):
    file_path = Path(file_path)

    blocks = extract_docx_blocks(
        file_path=file_path,
        evidence_id=evidence_id,
    )

    chunks = chunk_docx_blocks(blocks)

    extraction_status = (
        "extracted"
        if chunks
        else "no_extractable_text"
    )

    return {
        "evidence_id": evidence_id,
        "file_name": file_path.name,
        "file_type": ".docx",
        "block_count": len(blocks),
        "extraction_status": (
            extraction_status
        ),
        "chunk_count": len(chunks),
        "blocks": blocks,
        "chunks": chunks,
    }


# -------------------------
# XLSX
# -------------------------

def extract_xlsx_rows(
    file_path,
    evidence_id,
):
    file_path = Path(file_path)

    workbook = load_workbook(
        file_path,
        data_only=True,
        read_only=True,
    )

    rows = []

    for sheet_index, worksheet in enumerate(
        workbook.worksheets,
        start=1,
    ):
        for row_number, row in enumerate(
            worksheet.iter_rows(),
            start=1,
        ):
            values = []
            populated_cells = []

            for cell in row:
                if cell.value is None:
                    continue

                value = str(
                    cell.value
                ).strip()

                if not value:
                    continue

                values.append(value)
                populated_cells.append(
                    cell.coordinate
                )

            if not values:
                continue

            rows.append(
                {
                    "evidence_id": (
                        evidence_id
                    ),
                    "file_name": (
                        file_path.name
                    ),
                    "sheet_index": (
                        sheet_index
                    ),
                    "sheet_name": (
                        worksheet.title
                    ),
                    "row_number": (
                        row_number
                    ),
                    "cell_range": (
                        f"{populated_cells[0]}:"
                        f"{populated_cells[-1]}"
                    ),
                    "text": " | ".join(
                        values
                    ),
                }
            )

    workbook.close()

    return rows


def chunk_xlsx_rows(rows):
    chunks = []

    for row in rows:
        row_chunks = (
            split_text_into_chunks(
                row["text"]
            )
        )

        for chunk in row_chunks:
            source_id = (
                f"{row['evidence_id']}"
                f"-S{row['sheet_index']:02d}"
                f"-R{row['row_number']:04d}"
                f"-C{chunk['chunk_number']:02d}"
            )

            chunks.append(
                {
                    "source_id": (
                        source_id
                    ),
                    "evidence_id": row[
                        "evidence_id"
                    ],
                    "file_name": row[
                        "file_name"
                    ],
                    "file_type": ".xlsx",
                    "chunk_number": chunk[
                        "chunk_number"
                    ],
                    "text": chunk["text"],
                    "provenance": {
                        "sheet_name": row[
                            "sheet_name"
                        ],
                        "row_number": row[
                            "row_number"
                        ],
                        "cell_range": row[
                            "cell_range"
                        ],
                    },
                }
            )

    return chunks


def process_xlsx(
    file_path,
    evidence_id,
):
    file_path = Path(file_path)

    rows = extract_xlsx_rows(
        file_path=file_path,
        evidence_id=evidence_id,
    )

    chunks = chunk_xlsx_rows(rows)

    extraction_status = (
        "extracted"
        if chunks
        else "no_extractable_text"
    )

    return {
        "evidence_id": evidence_id,
        "file_name": file_path.name,
        "file_type": ".xlsx",
        "row_count": len(rows),
        "extraction_status": (
            extraction_status
        ),
        "chunk_count": len(chunks),
        "rows": rows,
        "chunks": chunks,
    }


# -------------------------
# CSV
# -------------------------

def extract_csv_rows(
    file_path,
    evidence_id,
):
    file_path = Path(file_path)

    rows = []

    with open(
        file_path,
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as csv_file:
        reader = csv.DictReader(
            csv_file
        )

        fieldnames = (
            reader.fieldnames or []
        )

        for row_number, row in enumerate(
            reader,
            start=2,
        ):
            values = []

            for column_name in fieldnames:
                value = row.get(
                    column_name
                )

                if value is None:
                    continue

                value = str(
                    value
                ).strip()

                if not value:
                    continue

                values.append(
                    f"{column_name}: "
                    f"{value}"
                )

            if not values:
                continue

            rows.append(
                {
                    "evidence_id": (
                        evidence_id
                    ),
                    "file_name": (
                        file_path.name
                    ),
                    "row_number": (
                        row_number
                    ),
                    "columns": (
                        fieldnames
                    ),
                    "text": " | ".join(
                        values
                    ),
                }
            )

    return rows


def chunk_csv_rows(rows):
    chunks = []

    for row in rows:
        row_chunks = (
            split_text_into_chunks(
                row["text"]
            )
        )

        for chunk in row_chunks:
            source_id = (
                f"{row['evidence_id']}"
                f"-R{row['row_number']:04d}"
                f"-C{chunk['chunk_number']:02d}"
            )

            chunks.append(
                {
                    "source_id": (
                        source_id
                    ),
                    "evidence_id": row[
                        "evidence_id"
                    ],
                    "file_name": row[
                        "file_name"
                    ],
                    "file_type": ".csv",
                    "chunk_number": chunk[
                        "chunk_number"
                    ],
                    "text": chunk["text"],
                    "provenance": {
                        "row_number": row[
                            "row_number"
                        ],
                        "columns": row[
                            "columns"
                        ],
                    },
                }
            )

    return chunks


def process_csv(
    file_path,
    evidence_id,
):
    file_path = Path(file_path)

    rows = extract_csv_rows(
        file_path=file_path,
        evidence_id=evidence_id,
    )

    chunks = chunk_csv_rows(rows)

    extraction_status = (
        "extracted"
        if chunks
        else "no_extractable_text"
    )

    return {
        "evidence_id": evidence_id,
        "file_name": file_path.name,
        "file_type": ".csv",
        "row_count": len(rows),
        "extraction_status": (
            extraction_status
        ),
        "chunk_count": len(chunks),
        "rows": rows,
        "chunks": chunks,
    }


# -------------------------
# TXT
# -------------------------

def extract_txt_lines(
    file_path,
    evidence_id,
):
    file_path = Path(file_path)

    with open(
        file_path,
        "r",
        encoding="utf-8-sig",
    ) as text_file:
        raw_lines = (
            text_file.readlines()
        )

    lines = []

    for line_number, line in enumerate(
        raw_lines,
        start=1,
    ):
        text = line.strip()

        if not text:
            continue

        lines.append(
            {
                "evidence_id": (
                    evidence_id
                ),
                "file_name": (
                    file_path.name
                ),
                "line_number": (
                    line_number
                ),
                "text": text,
            }
        )

    return lines


def chunk_txt_lines(
    lines,
    lines_per_block=20,
):
    chunks = []

    for start_index in range(
        0,
        len(lines),
        lines_per_block,
    ):
        block = lines[
            start_index:
            start_index + lines_per_block
        ]

        if not block:
            continue

        text = "\n".join(
            line["text"]
            for line in block
        )

        text_chunks = (
            split_text_into_chunks(
                text
            )
        )

        start_line = block[0][
            "line_number"
        ]

        end_line = block[-1][
            "line_number"
        ]

        for chunk in text_chunks:
            source_id = (
                f"{block[0]['evidence_id']}"
                f"-L{start_line:04d}"
                f"-L{end_line:04d}"
                f"-C{chunk['chunk_number']:02d}"
            )

            chunks.append(
                {
                    "source_id": (
                        source_id
                    ),
                    "evidence_id": block[
                        0
                    ]["evidence_id"],
                    "file_name": block[
                        0
                    ]["file_name"],
                    "file_type": ".txt",
                    "chunk_number": chunk[
                        "chunk_number"
                    ],
                    "text": chunk["text"],
                    "provenance": {
                        "start_line": (
                            start_line
                        ),
                        "end_line": (
                            end_line
                        ),
                    },
                }
            )

    return chunks


def process_txt(
    file_path,
    evidence_id,
):
    file_path = Path(file_path)

    lines = extract_txt_lines(
        file_path=file_path,
        evidence_id=evidence_id,
    )

    chunks = chunk_txt_lines(lines)

    extraction_status = (
        "extracted"
        if chunks
        else "no_extractable_text"
    )

    return {
        "evidence_id": evidence_id,
        "file_name": file_path.name,
        "file_type": ".txt",
        "line_count": len(lines),
        "extraction_status": (
            extraction_status
        ),
        "chunk_count": len(chunks),
        "lines": lines,
        "chunks": chunks,
    }


# -------------------------
# Dispatcher
# -------------------------

def process_evidence_files(
    file_paths,
):
    evidence_documents = []

    for index, file_path in enumerate(
        file_paths,
        start=1,
    ):
        file_path = Path(file_path)

        evidence_id = (
            f"E-{index:03d}"
        )

        file_type = (
            file_path.suffix.lower()
        )

        if file_type == ".pdf":
            document = process_pdf(
                file_path=file_path,
                evidence_id=evidence_id,
            )

        elif file_type == ".docx":
            document = process_docx(
                file_path=file_path,
                evidence_id=evidence_id,
            )

        elif file_type == ".xlsx":
            document = process_xlsx(
                file_path=file_path,
                evidence_id=evidence_id,
            )

        elif file_type == ".csv":
            document = process_csv(
                file_path=file_path,
                evidence_id=evidence_id,
            )

        elif file_type == ".txt":
            document = process_txt(
                file_path=file_path,
                evidence_id=evidence_id,
            )

        else:
            document = {
                "evidence_id": (
                    evidence_id
                ),
                "file_name": (
                    file_path.name
                ),
                "file_type": (
                    file_type
                ),
                "extraction_status": (
                    "unsupported"
                ),
                "chunk_count": 0,
                "chunks": [],
            }

        evidence_documents.append(
            document
        )

    return evidence_documents


def extract_pdf_evidence(
    file_paths,
):
    """
    Backward-compatible helper
    from Day 3.
    """

    return [
        process_pdf(
            file_path=file_path,
            evidence_id=f"E-{index:03d}",
        )
        for index, file_path in enumerate(
            file_paths,
            start=1,
        )
    ]